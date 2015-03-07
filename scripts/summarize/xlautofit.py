#Copyright [2014] [Puget Sound Regional Council]

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

#    http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.

import xlrd
import sys, os
sys.path.append(os.path.join(os.getcwd(),"scripts"))

def run(file): #This function "autofits" all of the columns in the workbook. However, it removes all charts and images
    print('WARNING: xlautofit.run() removes charts, images, and styles')
    import openpyxl
    wd={}
    colnums=[]
    book=xlrd.open_workbook(file, on_demand = True)
    sheetnames = book.sheet_names()
    for sheet in sheetnames:
        ws = book.sheet_by_name(sheet)
        colwidths = []
        cols = ws.ncols
        for colnum in range(cols):
            widths = []
            for rownum in range(len(ws.col(colnum))):
                widths.append(len(str(ws.cell(rownum,colnum).value)))
            colwidths.append(max(widths))
        wd.update({sheet: colwidths})
        colnums.append(cols)
    wbook = openpyxl.load_workbook(file)
    for sheetnum in range(len(wbook.get_sheet_names())):
        wsheet = wbook.get_sheet_by_name(wbook.get_sheet_names()[sheetnum])
        for colnum in range(colnums[sheetnum]):
            wsheet.column_dimensions[openpyxl.cell.get_column_letter(colnum+1)].width = wd[sheetnames[sheetnum]][colnum]
    wbook.save(file)

def getwidths(file): #This function returns a dictionary where the keys are the sheet names and the values are lists of the necessary column widths for the sheets
    wd = {}
    book = xlrd.open_workbook(file, on_demand = True)
    sheetnames = book.sheet_names()
    for sheet in sheetnames:
        ws = book.sheet_by_name(sheet)
        colwidths = []
        cols = ws.ncols
        for colnum in range(cols):
            widths = []
            for rownum in range(len(ws.col(colnum))):
                widths.append(len(str(ws.cell(rownum,colnum).value)))
            colwidths.append(max(widths))
        wd.update({sheet: colwidths})
    return(wd)

def getmaxwidths(file): #Like the last one, but every non-index or filler column is the same width
    wd = {}
    book = xlrd.open_workbook(file, on_demand = True)
    sheetnames = book.sheet_names()
    for sheet in sheetnames:
        ws = book.sheet_by_name(sheet)
        colwidths = [] #This list is the necessary widths for all columns
        ni_widths = [] #This one is for columns that aren't indices or fillers
        cols = ws.ncols
        for colnum in range(cols):
            widths = []
            for rownum in range(len(ws.col(colnum))):
                widths.append(len(str(ws.cell(rownum,colnum).value)))
            colwidths.append(max(widths))
            if colnum > 0:
                if colwidths[colnum - 1] > 1.5 and colwidths[colnum] > 1.5:
                    ni_widths.append(colwidths[colnum])
        non_index_width = max(ni_widths)
        for colnum in range(1, cols): #This loop makes all non-index or filler columns the same width (as wide as the widest one)
            if colwidths[colnum] > 1.5 and colwidths[colnum - 1] > 1.5:
                colwidths[colnum] = non_index_width
        wd.update({sheet: colwidths})
    return(wd)

def even_widths_single_index(file):
    wd = {}
    book = xlrd.open_workbook(file, on_demand = True)
    sheetnames = book.sheet_names()
    for sheet in sheetnames:
        ws = book.sheet_by_name(sheet)
        colwidths = [] #This list is the necessary widths for all columns
        ni_widths = [] #This one is for columns that aren't indices or fillers
        cols = ws.ncols
        for colnum in range(cols):
            widths = []
            for rownum in range(len(ws.col(colnum))):
                widths.append(len(str(ws.cell(rownum,colnum).value)))
            colwidths.append(max(widths))
            if colnum > 0:
                if colwidths[colnum] > 1.5:
                    ni_widths.append(colwidths[colnum])
        non_index_width = max(ni_widths)
        for colnum in range(1, cols): #This loop makes all non-index or filler columns the same width (as wide as the widest one)
            if colwidths[colnum] > 1.5:
                colwidths[colnum] = non_index_width
        wd.update({sheet: colwidths})
    return(wd)